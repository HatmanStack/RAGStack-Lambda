"""
XLSX extractor using openpyxl.

Extracts Excel workbook content with multi-sheet awareness,
similar to CSV smart extraction for each sheet.
"""

import io
from datetime import datetime
from typing import Any

from openpyxl import load_workbook

from .base import BaseExtractor, ExtractionResult


class XlsxExtractor(BaseExtractor):
    """Extract content from XLSX files.

    Features:
    - Extract multiple sheets with CSV-style smart extraction
    - Handle various cell types (text, number, date, formula)
    - Extract workbook properties
    """

    MAX_SAMPLE_ROWS = 5

    def extract(self, content: bytes, filename: str) -> ExtractionResult:
        """Extract text content from XLSX file bytes.

        Args:
            content: Raw XLSX content as bytes.
            filename: Original filename.

        Returns:
            ExtractionResult with markdown content and metadata.
        """
        title = self._extract_title_from_filename(filename)

        # Read XLSX
        try:
            wb = load_workbook(io.BytesIO(content), data_only=True)
        except Exception as e:
            # Fall back with warning
            return self._create_fallback_result(content, filename, title, str(e))

        # Extract metadata
        sheet_names = wb.sheetnames
        sheet_count = len(sheet_names)

        # Extract workbook properties
        created = None
        modified = None
        if wb.properties:
            created = wb.properties.created
            modified = wb.properties.modified

        # Calculate total rows across all sheets
        total_rows = 0
        sheet_data: list[tuple[str, list[list[Any]]]] = []

        for sheet_name in sheet_names:
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                # Convert row to list, handling None values
                row_data = [self._format_cell(cell) for cell in row]
                if any(cell for cell in row_data):  # Skip empty rows
                    rows.append(row_data)

            if rows:
                total_rows += len(rows) - 1  # Exclude header
                sheet_data.append((sheet_name, rows))

        # Build structural metadata
        structural_metadata: dict[str, Any] = {
            "sheet_count": sheet_count,
            "sheets": sheet_names,
            "total_rows": total_rows,
        }
        if created:
            structural_metadata["created"] = created.isoformat() if isinstance(created, datetime) else str(created)
        if modified:
            structural_metadata["modified"] = modified.isoformat() if isinstance(modified, datetime) else str(modified)

        # Generate markdown
        markdown_body = self._generate_markdown(filename, sheet_data)
        word_count = self._count_words(markdown_body)

        # Build frontmatter metadata
        frontmatter_metadata = {
            "source_file": filename,
            "file_type": "xlsx",
            "sheet_count": sheet_count,
            "sheets": sheet_names,
            "total_rows": total_rows,
        }

        frontmatter = self._generate_frontmatter(frontmatter_metadata)

        # Combine frontmatter and content
        markdown = f"{frontmatter}\n{markdown_body}"

        return ExtractionResult(
            markdown=markdown,
            file_type="xlsx",
            title=title,
            word_count=word_count,
            structural_metadata=structural_metadata,
            parse_warning=None,
        )

    def _format_cell(self, cell: Any) -> str:
        """Format cell value as string."""
        if cell is None:
            return ""
        if isinstance(cell, datetime):
            return cell.strftime("%Y-%m-%d")
        if isinstance(cell, bool):
            return "Yes" if cell else "No"
        if isinstance(cell, float):
            # Format floats nicely
            if cell == int(cell):
                return str(int(cell))
            return f"{cell:.2f}"
        return str(cell)

    def _generate_markdown(
        self,
        filename: str,
        sheet_data: list[tuple[str, list[list[Any]]]],
    ) -> str:
        """Generate markdown from XLSX data."""
        lines = []

        # Title
        lines.append(f"# Workbook: {filename}")
        lines.append("")
        lines.append(f"Workbook with {len(sheet_data)} sheet{'s' if len(sheet_data) != 1 else ''}.")
        lines.append("")

        # Each sheet
        for sheet_name, rows in sheet_data:
            lines.append("---")
            lines.append("")
            lines.append(f"## Sheet: {sheet_name}")
            lines.append("")

            if not rows:
                lines.append("*Empty sheet*")
                lines.append("")
                continue

            # Get headers (first row)
            headers = rows[0]
            data_rows = rows[1:]

            lines.append(f"{len(data_rows)} rows, {len(headers)} columns.")
            lines.append("")

            # Column descriptions
            if data_rows:
                lines.append("### Columns")
                lines.append("")
                for i, header in enumerate(headers):
                    if not header:
                        header = f"Column{i + 1}"
                    # Get sample values
                    samples = []
                    for row in data_rows[:3]:
                        if i < len(row) and row[i]:
                            samples.append(str(row[i]))
                    sample_str = ", ".join(f'"{s}"' for s in samples[:3]) if samples else ""
                    lines.append(f"- **{header}**{': ' + sample_str if sample_str else ''}")
                lines.append("")

            # Sample data table
            lines.append("### Sample Data")
            lines.append("")

            # Table header
            header_strs = [h if h else f"Col{i + 1}" for i, h in enumerate(headers)]
            lines.append("| " + " | ".join(header_strs) + " |")
            lines.append("| " + " | ".join(["---"] * len(headers)) + " |")

            # Table rows (first N)
            for row in data_rows[: self.MAX_SAMPLE_ROWS]:
                # Pad row if needed
                padded_row = list(row) + [""] * (len(headers) - len(row))
                cells = [str(cell)[:50] + "..." if len(str(cell)) > 50 else str(cell) for cell in padded_row[: len(headers)]]
                lines.append("| " + " | ".join(cells) + " |")

            if len(data_rows) > self.MAX_SAMPLE_ROWS:
                lines.append(f"\n*...and {len(data_rows) - self.MAX_SAMPLE_ROWS} more rows*")

            lines.append("")

        return "\n".join(lines)

    def _create_fallback_result(
        self, content: bytes, filename: str, title: str, error: str
    ) -> ExtractionResult:
        """Create fallback result when XLSX parsing fails."""
        word_count = 0
        frontmatter_metadata = {
            "source_file": filename,
            "file_type": "xlsx",
            "parse_warning": f"XLSX parsing failed: {error}",
        }
        frontmatter = self._generate_frontmatter(frontmatter_metadata)
        markdown = f"{frontmatter}\nFailed to parse XLSX: {error}"

        return ExtractionResult(
            markdown=markdown,
            file_type="xlsx",
            title=title,
            word_count=word_count,
            structural_metadata={},
            parse_warning=f"XLSX parsing failed: {error}",
        )
